import pandas as pd
from datetime import datetime, timedelta, date
import shutil
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl import load_workbook
import os
import glob
import numpy as np
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import time

def createpath():
    CAFIFfilepath=(os.path.join(r"C:\Users\odemirhan\TURISTIK HAVA TASIMACILIK A.S\Gökmen Düzgören - FOE_2019\phyton\db_python\OCC\Sharefolder", tomorrowasstr))
    if not os.path.exists(CAFIFfilepath):
        os.mkdir(CAFIFfilepath)
    
dayy=1  #today icin 0 tomorrow için 1
today=datetime.strptime(datetime.strftime(datetime.now(), '%Y-%m-%d'), '%Y-%m-%d')

tomorrowP1=datetime.strptime(datetime.strftime(datetime.now() + timedelta(days=2), '%Y-%m-%d'), '%Y-%m-%d')
tomorrow=datetime.strptime(datetime.strftime(datetime.now() + timedelta(days=dayy), '%Y-%m-%d'), '%Y-%m-%d')
tomorrowmonstr=datetime.strftime(datetime.now() + timedelta(days=dayy), '%b')
tomorrowmonstr=tomorrowmonstr.upper()
tomorrowdaystr=datetime.strftime(datetime.now() + timedelta(days=dayy), '%d')

tomorrowasstr=datetime.strftime(datetime.now() + timedelta(days=dayy), '%Y-%m-%d')


for daycnt in range(2,5):
    try:
        todaymtwo=datetime.strftime(datetime.now()- timedelta(days=daycnt), '%Y-%m-%d')
        oldfilepaths=(os.path.join(r"C:\Users\odemirhan\TURISTIK HAVA TASIMACILIK A.S\Gökmen Düzgören - FOE_2019\phyton\db_python\OCC\Sharefolder", todaymtwo))
        if os.path.exists(oldfilepaths):
            shutil.rmtree(oldfilepaths)
    except:
        pass


SCHfile=r'C:\Users\odemirhan\TURISTIK HAVA TASIMACILIK A.S\coreSafety - VERSION_03/SCH_dynamic.csv'
OCCfile='\\192.168.2.4\occ\Dispatch_ORTAK\JETPLANNER IMPORT/'
DBstr='\\\\192.168.2.4\Muhendislik\InforCognosOutputs/'

MELDF=pd.read_excel('\\\\192.168.2.4\Muhendislik\InforCognosOutputs/FLEET_OPEN_HOLD_ITEM_LIST-en.xlsx')
MELDF=MELDF.dropna(subset=['AIRCRAFT'])
MELDF=MELDF[MELDF['MEL/AMM/CDL REF'].str.contains("MEL") | MELDF['MEL/AMM/CDL REF'].str.contains("CDL")]     
MELDF['DISCREPANCY']=MELDF['DISCREPANCY'].str.replace("DISCREPANCY:", "")
MELDF['WRITE2CAFIF']=MELDF['MEL/AMM/CDL REF']+":  "+MELDF['DISCREPANCY']
MELDF['WRITE2CAFIF'] = MELDF['WRITE2CAFIF'].str.rsplit('ACTION TAKEN', n=1).str.get(0)
MELDF=MELDF.reset_index(drop=True)


EFBDF=pd.read_excel(r'C:\Users\odemirhan\TURISTIK HAVA TASIMACILIK A.S\Gökmen Düzgören - FOE_2019\phyton\db_python\OCC\Sharefolder/EFB DB.xlsx', header=None)



OCCfiles = glob.glob('\\\\192.168.2.4\occ\Dispatch_ORTAK\JETPLANNER IMPORT/*.txt')
for cntfile in range(len(OCCfiles)):
    filenametxt=os.path.basename(OCCfiles[cntfile])
    filenametxt=filenametxt.upper()
    if tomorrowmonstr in filenametxt:
        if tomorrowdaystr in filenametxt:
            
            OCClatest=OCCfiles[cntfile]
            


SCHdf=pd.read_csv(SCHfile, low_memory=False)
OCCdf=pd.read_csv(OCClatest, header=None, low_memory=False,  dtype={10 : object, 1: object })

OCCdf["Date1"]=pd.to_datetime(OCCdf.iloc[:,14], format='%d%b%y')
OCCdf["Time"]=pd.to_datetime(OCCdf.iloc[:,10], format='%H%M').dt.time
OCCdf["Datetime"]=OCCdf["Date1"].astype(str)+" "+ OCCdf["Time"].astype(str)


OCCdf["Reg"]=OCCdf.iloc[:,2]
OCCdf["Reg"]=OCCdf["Reg"].str[1:6]#+"-"+OCCdf["Reg"].str[3:6]
OCCdf['OCCkey']=OCCdf["Date1"].astype(str)+OCCdf.iloc[:,4]+OCCdf.iloc[:,5]+OCCdf.iloc[:,8]
#+OCCdf["Reg"]

SCHdf['Date']=SCHdf['STD'].str[0:10]
SCHdf['STD']=pd.to_datetime(SCHdf['STD'])

SCHdf=SCHdf[SCHdf['STD']>tomorrow]
SCHdf=SCHdf[SCHdf['STD']<tomorrowP1]
SCHdf=SCHdf.reset_index()
SCHdf['SCHkey']=SCHdf['Date']+SCHdf["SDEP"]+SCHdf["SARR"]+SCHdf["CARRIER"]+SCHdf["FNO"].astype(str)
#+SCHdf["ACREG"].str[0:5]
FinalDF=OCCdf.merge(SCHdf, left_on='OCCkey', right_on='SCHkey')
FinalDF=FinalDF[["Date", 8,1, "Reg",3,4,5,"C1","C2","C3"]]

createpath()
FinalDF["CrewKey"]=FinalDF["C1"].str[:3]+"-"+FinalDF["C2"].str[:3]+"-"+FinalDF["C3"].fillna("")
FinalDF=FinalDF.drop_duplicates()
FinalDF=FinalDF.sort_values(by=[1])


FinalDF[1]=FinalDF[1].astype(str)
FinalDF=FinalDF.reset_index(drop=True)
subFinalDF=FinalDF

cntfl=0
while cntfl<len(FinalDF):

    if len(FinalDF[FinalDF["CrewKey"]==FinalDF.at[cntfl, "CrewKey"]])>1:
        dummyFinalDF=FinalDF[FinalDF["CrewKey"]==FinalDF.at[cntfl, "CrewKey"]]
        dummyFinalDF=dummyFinalDF.reset_index(drop=True)
        dateasstr=FinalDF.iat[cntfl,0]
        Date2wrt=datetime.strftime(datetime.strptime(dateasstr, '%Y-%m-%d'), '%d.%m.%Y')
        Reg2wrt=(FinalDF.iat[cntfl,3])[0:2]+"-"+(FinalDF.iat[cntfl,3])[2:5]

        try:
            Cpt2wrt=(FinalDF.iat[cntfl,7])[0:3]
        except:
            Cpt2wrt=""
        try:
            FO2wrt=(FinalDF.iat[cntfl,8])[0:3]
        except:
            FO2wrt=""
            
        
        if pd.isna(FinalDF.iat[cntfl,9]):
            TFO2wrt=""
        else:
            TFO2wrt=(FinalDF.iat[cntfl,9])[0:3]
        FLno=[0] * len(dummyFinalDF)
        dummytime=[0] * len(dummyFinalDF)
        Time2wrt=[0] * len(dummyFinalDF)
        Dep2wrt=[0] * len(dummyFinalDF)
        Arr2wrt=[0] * len(dummyFinalDF)
        EnrouteWarning=[0] * len(dummyFinalDF)
        DWE=[0] * len(dummyFinalDF)
        Pax2wrt=[0] * len(dummyFinalDF)
        for cntfl1 in range(len(dummyFinalDF)):   
            FLno[cntfl1]=dummyFinalDF.iat[cntfl1,1]
            
            dtime=dummyFinalDF.iat[cntfl1,2]
            
                
            dummytime[cntfl1]=dtime    
            Time2wrt[cntfl1]=dtime[0:2] +":"+dtime[2:4] +" UTC"



        
            Pax2wrt[cntfl1]=str(dummyFinalDF.iat[cntfl1,4])
            Dep2wrt[cntfl1]=dummyFinalDF.iat[cntfl1,5]
            Arr2wrt[cntfl1]=dummyFinalDF.iat[cntfl1,6]
            EnrouteWarning[cntfl1]=str(FLno[cntfl1])+" WEATHER PERMISSIBLE AERODROMES BASED ON 394 NMs CIRCLES ARE:"
            DWE[cntfl1]=str(FLno[cntfl1])+" "+str(Dep2wrt[cntfl1])+"-"+ str(Arr2wrt[cntfl1])+" Dispatch Weather extra fuel:"
        
        CAFIFname="-".join(FLno)
        #for cnt3 in range(len(FLno)):
         #   CAFIFname=CAFIFname+"-"+FLno[cnt3]

        copypathCAI=r"C:\Users\odemirhan\TURISTIK HAVA TASIMACILIK A.S\Gökmen Düzgören - FOE_2019\phyton\db_python\OCC\Sharefolder/"+tomorrowasstr+"/"+dummytime[0]+"-"+CAFIFname+"_CAFIF.xlsx"

        try:
            shutil.copy(r"C:\Users\odemirhan\TURISTIK HAVA TASIMACILIK A.S\Gökmen Düzgören - FOE_2019\phyton\db_python\OCC\Org/CAI CAFIF.xlsx", copypathCAI)
            time.sleep(2)
        except:
            pass

        time.sleep(2)

        workbook   = load_workbook(filename=copypathCAI)
        ws = workbook.active
        dummyexcellist=["F","J","N","R","V"]
        ws['F7']=Cpt2wrt+" "+FO2wrt+" "+TFO2wrt+"+ X CABIN CREW"
        ws['F10']=Date2wrt

        for cnt2 in range(len(dummyFinalDF)):
            
            ws[dummyexcellist[cnt2]+"11"]=Dep2wrt[cnt2]+" - "+Arr2wrt[cnt2]
            ws[dummyexcellist[cnt2]+"12"]=Time2wrt[cnt2]
            ws[dummyexcellist[cnt2]+"14"]=FLno[cnt2]
        
        
        ws['F16']=Reg2wrt

        ws['K22']=" - ".join(Pax2wrt)
        if Reg2wrt[0:2]=="9H":
            ws['R1']=""
            ws['A47']="’Please don’t forget to submit your fatigue status through the TOD survey’"
            currentCell = ws['A47']
            currentCell.alignment = Alignment(horizontal='center')
            currentCell.font=Font(size=10, bold=True)
        else:
            ws['A45']="Check for the presence and validity of Medical Certificate, crew license and appropriate ratings"
            ws['A46']="Check for the presence spare correcting spectacles (in case a flight crew member is required to wear corrective lenses and glasses)"
            ws['A47']="Check for the crew composition is matching(OML)"																							

           
            for cntdummy in range(3):
                currentCell = ws['A'+str(45+cntdummy)]
                currentCell.alignment = Alignment(horizontal='center')
                currentCell.font=Font(size=10, bold=False)
                


        ws['K24']="OK"
        
        ws['K28']="NIL"

        dummyMELDF=MELDF[MELDF['AIRCRAFT']==Reg2wrt]

        if len(dummyMELDF)==0:
            ws['E31']="NIL"
            currentCell = ws['E31']
            currentCell.alignment = Alignment(horizontal='center')
            currentCell.font=Font(size=13)

        elif len(dummyMELDF)<5:
            for cntMEL in range(len(dummyMELDF)):
                cntmel2=str(30+cntMEL)
                ws['E'+cntmel2]=dummyMELDF.iat[cntMEL, 13]
                currentCell = ws['E'+cntmel2]
                currentCell.alignment = Alignment(horizontal='center')
        else:
            ws['E30']="THERE ARE HOLD ITEMS MORE THAN 4. PLEASE TYPE IT MANUALLY"


        for cntDWE in range(len(EnrouteWarning)):
            ws['A'+str(36+cntDWE)]=EnrouteWarning[cntDWE]
            ws['A'+str(39+cntDWE)]=DWE[cntDWE]
            
            currentCell = ws['A'+str(36+cntDWE)]
            currentCell.alignment = Alignment(horizontal='center')
            currentCell.font=Font(size=10, bold=False)
            currentCell = ws['A'+str(39+cntDWE)]
            currentCell.alignment = Alignment(horizontal='center')
            currentCell.font=Font(size=10, bold=False)

        

        for cntEFB in range(len(EFBDF)):
            cntEFB2=str(42+cntEFB)
            ws['A'+cntEFB2]=EFBDF.iat[cntEFB, 0]
            currentCell = ws['A'+cntEFB2]
            currentCell.alignment = Alignment(horizontal='center')
            currentCell.font=Font(size=10, bold=False)
        
            
        try:
            workbook.save(filename=copypathCAI)
        except:
            try:
                time.sleep(2)
                workbook.save(filename=copypathCAI)
            except:
                pass
            
        
        FinalDF=pd.concat([dummyFinalDF,FinalDF]).drop_duplicates(keep=False)
        FinalDF=FinalDF.reset_index(drop=True)
        
    else:
        
        

        dateasstr=FinalDF.iat[cntfl,0]
        Date2wrt=datetime.strftime(datetime.strptime(dateasstr, '%Y-%m-%d'), '%d.%m.%Y')
        FLno=FinalDF.iat[cntfl,1]
        dtime=FinalDF.iat[cntfl, 2]
        Time2wrt=dtime[0:2] +":"+dtime[2:4] +" UTC"
        Reg2wrt=(FinalDF.iat[cntfl,3])[0:2]+"-"+(FinalDF.iat[cntfl,3])[2:5]
        Pax2wrt=str(FinalDF.iat[cntfl,4])
        Dep2wrt=FinalDF.iat[cntfl,5]
        Arr2wrt=FinalDF.iat[cntfl,6]
        try:
            Cpt2wrt=(FinalDF.iat[cntfl,7])[0:3]
        except:
            Cpt2wrt=""
        try:
            FO2wrt=(FinalDF.iat[cntfl,8])[0:3]
        except:
            FO2wrt=""
            
        
            
        if pd.isna(FinalDF.iat[cntfl,9]):
            TFO2wrt=""
        else:
            TFO2wrt=(FinalDF.iat[cntfl,9])[0:3]

        copypathCAI=r"C:\Users\odemirhan\TURISTIK HAVA TASIMACILIK A.S\Gökmen Düzgören - FOE_2019\phyton\db_python\OCC\Sharefolder/"+tomorrowasstr+"/"+dtime+"-"+FLno+"_CAFIF.xlsx"
        try:
            
            shutil.copy(r"C:\Users\odemirhan\TURISTIK HAVA TASIMACILIK A.S\Gökmen Düzgören - FOE_2019\phyton\db_python\OCC\Org/CAI CAFIF.xlsx", copypathCAI)
            time.sleep(2)
        except:
            pass


        workbook   = load_workbook(filename=copypathCAI)
        ws = workbook.active

        ws['F7']=Cpt2wrt+" "+FO2wrt+" "+TFO2wrt+"+ X CABIN CREW"
        ws['F10']=Date2wrt
        ws['F11']=Dep2wrt+" - "+Arr2wrt
        ws['F12']=Time2wrt
        ws['F14']=FLno
        ws['F16']=Reg2wrt
        ws['K22']=Pax2wrt
        
        ws['A36']=str(FLno)+ " WEATHER PERMISSIBLE AERODROMES BASED ON 394 NMs CIRCLES ARE:"
        ws['A39']=str(FLno)+ Dep2wrt+"-"+Arr2wrt+" Dispatch Weather extra fuel:"
            
            
        currentCell = ws['A36']
        currentCell.alignment = Alignment(horizontal='center')
        currentCell.font=Font(size=10, bold=False)
        currentCell = ws['A39']
        currentCell.alignment = Alignment(horizontal='center')
        currentCell.font=Font(size=10, bold=False)
        if Reg2wrt[0:2]=="9H":
            ws['R1']=""
            ws['A47']="’Please don’t forget to submit your fatigue status through the TOD survey’"
            currentCell = ws['A47']
            currentCell.alignment = Alignment(horizontal='center')
            currentCell.font=Font(size=10, bold=True)
        else:
            ws['A45']="Check for the presence and validity of Medical Certificate, crew license and appropriate ratings"
            ws['A46']="Check for the presence spare correcting spectacles (in case a flight crew member is required to wear corrective lenses and glasses)"
            ws['A47']="Check for the crew composition is matching(OML)"																							

           
            for cntdummy in range(3):
                currentCell = ws['A'+str(45+cntdummy)]
                currentCell.alignment = Alignment(horizontal='center')
                currentCell.font=Font(size=10, bold=False)
            

        ws['K24']="OK"
       
        ws['K28']="NIL"

        dummyMELDF=MELDF[MELDF['AIRCRAFT']==Reg2wrt]

        if len(dummyMELDF)==0:
            ws['E31']="NIL"
            currentCell = ws['E31']
            currentCell.alignment = Alignment(horizontal='center')
            currentCell.font=Font(size=13)
            

        elif len(dummyMELDF)<5:
            for cntMEL in range(len(dummyMELDF)):
                cntmel2=str(30+cntMEL)
                ws['E'+cntmel2]=dummyMELDF.iat[cntMEL, 13]
                currentCell = ws['E'+cntmel2]
                currentCell.alignment = Alignment(horizontal='center')
        else:
            ws['E30']="THERE ARE HOLD ITEMS MORE THAN 4. PLEASE TYPE IT MANUALLY"



        for cntEFB in range(len(EFBDF)):
            cntEFB2=str(41+cntEFB)
            ws['A'+cntEFB2]=EFBDF.iat[cntEFB, 0]
            currentCell = ws['A'+cntEFB2]
            currentCell.alignment = Alignment(horizontal='center')
            currentCell.font=Font(size=10, bold=False)

        try:
            workbook.save(filename=copypathCAI)
        except:
            try:
                time.sleep(2)
                workbook.save(filename=copypathCAI)
            except:
                pass
        cntfl+=1
